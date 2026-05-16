"""PlaceUp Career H1B sponsor importer.

Imports curated H1B source files into the `h1b_sponsors` table. The import is
idempotent: re-running updates existing sponsor rows in place.
"""

from __future__ import annotations

import csv
import hashlib
import logging
from pathlib import Path
from typing import Iterable

logger = logging.getLogger(__name__)

BACKEND_DIR = Path(__file__).resolve().parents[2]
EXCEL_FILE = BACKEND_DIR / "H1b_US_DataLIst.xlsx"
H1B_DATA_DIR = BACKEND_DIR / "data" / "h1b"
COMPANIES_CSV = H1B_DATA_DIR / "h1b_us_companies.csv"
SPONSORS_CSV = H1B_DATA_DIR / "sponsors_2024_2025_2026.csv"


def _safe_int(value) -> int:
    try:
        raw = str(value or "").replace(",", "").strip()
        return int(float(raw)) if raw else 0
    except Exception:
        return 0


def _safe_text(value) -> str:
    return str(value or "").strip()


def _sponsor_id(employer: str, city: str, state: str, fiscal_year: int) -> str:
    seed = f"{employer.lower()}|{city.lower()}|{state.lower()}|{fiscal_year}"
    return hashlib.sha1(seed.encode("utf-8")).hexdigest()[:16]


def _row_to_sponsor(row: tuple) -> dict | None:
    """Pluck the columns we care about from the legacy Excel source row."""
    try:
        employer = _safe_text(row[2])
        if not employer:
            return None

        fiscal_year = _safe_int(row[1])
        city = _safe_text(row[5])
        state = _safe_text(row[6])
        zip_code = _safe_text(row[7])

        new_a = _safe_int(row[8])
        new_d = _safe_int(row[9])
        cont_a = _safe_int(row[10])
        cont_d = _safe_int(row[11])
        chg_a = _safe_int(row[12]) + _safe_int(row[14]) + _safe_int(row[16]) + _safe_int(row[18])
        chg_d = _safe_int(row[13]) + _safe_int(row[15]) + _safe_int(row[17]) + _safe_int(row[19])

        total = new_a + new_d + cont_a + cont_d + chg_a + chg_d
        if total == 0:
            return None

        return {
            "id": _sponsor_id(employer, city, state, fiscal_year),
            "employer_name": employer,
            "city": city,
            "state": state,
            "zip_code": zip_code,
            "initial_approvals": new_a,
            "initial_denials": new_d,
            "continuing_approvals": cont_a + chg_a,
            "continuing_denials": cont_d + chg_d,
            "total_petitions": total,
            "fiscal_year": fiscal_year,
            "data_json": {"source_file": EXCEL_FILE.name},
        }
    except Exception:
        return None


def _company_csv_row_to_sponsor(row: dict) -> dict | None:
    employer = _safe_text(row.get("employer_name"))
    if not employer:
        return None

    fiscal_year = _safe_int(row.get("fiscal_year"))
    city = _safe_text(row.get("city"))
    state = _safe_text(row.get("state"))
    approvals = _safe_int(row.get("approvals"))
    denials = _safe_int(row.get("denials"))
    petitions = _safe_int(row.get("petitions")) or approvals + denials
    if petitions == 0:
        return None

    return {
        "id": _sponsor_id(employer, city, state, fiscal_year),
        "employer_name": employer,
        "city": city,
        "state": state,
        "zip_code": "",
        "initial_approvals": approvals,
        "initial_denials": denials,
        "continuing_approvals": 0,
        "continuing_denials": 0,
        "total_petitions": petitions,
        "fiscal_year": fiscal_year,
        "data_json": {"industry": _safe_text(row.get("industry")), "source_file": COMPANIES_CSV.name},
    }


def _sponsors_csv_row_to_sponsor(row: dict) -> dict | None:
    employer = _safe_text(row.get("employer_name"))
    if not employer:
        return None

    fiscal_year = _safe_int(row.get("fiscal_year"))
    city = _safe_text(row.get("city"))
    state = _safe_text(row.get("state"))
    initial_approvals = _safe_int(row.get("initial_approvals"))
    initial_denials = _safe_int(row.get("initial_denials"))
    continuing_approvals = _safe_int(row.get("continuing_approvals"))
    continuing_denials = _safe_int(row.get("continuing_denials"))
    total_petitions = _safe_int(row.get("total_petitions")) or (
        initial_approvals + initial_denials + continuing_approvals + continuing_denials
    )
    if total_petitions == 0:
        return None

    return {
        "id": _safe_text(row.get("id"))[:64] or _sponsor_id(employer, city, state, fiscal_year),
        "employer_name": employer,
        "city": city,
        "state": state,
        "zip_code": _safe_text(row.get("zip_code")),
        "initial_approvals": initial_approvals,
        "initial_denials": initial_denials,
        "continuing_approvals": continuing_approvals,
        "continuing_denials": continuing_denials,
        "total_petitions": total_petitions,
        "fiscal_year": fiscal_year,
        "data_json": {"source_file": SPONSORS_CSV.name},
    }


def _iter_excel_rows(path: Path) -> Iterable[tuple]:
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    first = True
    for row in ws.iter_rows(values_only=True):
        if first:
            first = False
            continue
        yield row


def _iter_csv_sponsors(path: Path) -> Iterable[dict]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            sponsor = _company_csv_row_to_sponsor(row) if path.name == COMPANIES_CSV.name else _sponsors_csv_row_to_sponsor(row)
            if sponsor:
                yield sponsor


async def import_h1b_excel(db, *, file_path: Path = EXCEL_FILE, force: bool = False) -> int:
    """Import H1B sponsor data into the configured database."""
    if not force:
        try:
            existing = await db.get_h1b_sponsors(limit=1)
            if existing:
                logger.info("H1B sponsors already populated; skipping import (force=False)")
                return 0
        except Exception:
            pass

    aggregated: dict[str, dict] = {}
    csv_paths = [path for path in (COMPANIES_CSV, SPONSORS_CSV) if path.exists()]

    if csv_paths:
        logger.info("H1B: starting CSV import from %s", ", ".join(path.name for path in csv_paths))
        for path in csv_paths:
            for sponsor in _iter_csv_sponsors(path):
                aggregated[sponsor["id"]] = sponsor
    elif file_path.exists():
        logger.info("H1B: starting Excel import from %s", file_path.name)
        for row in _iter_excel_rows(file_path):
            sponsor = _row_to_sponsor(row)
            if sponsor is None:
                continue
            sid = sponsor["id"]
            if sid in aggregated:
                agg = aggregated[sid]
                for key in (
                    "initial_approvals",
                    "initial_denials",
                    "continuing_approvals",
                    "continuing_denials",
                    "total_petitions",
                ):
                    agg[key] += sponsor[key]
            else:
                aggregated[sid] = sponsor
    else:
        logger.warning("H1B source files not found under %s; skipping import", H1B_DATA_DIR)
        return 0

    sponsors = list(aggregated.values())
    if not sponsors:
        logger.warning("H1B: no usable sponsor rows extracted")
        return 0

    written = await db.upsert_h1b_sponsors(sponsors)
    logger.info("H1B: imported %s sponsor records (%s unique IDs)", written, len(sponsors))
    return written
