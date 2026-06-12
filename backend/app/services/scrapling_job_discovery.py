"""Scrapling-powered job discovery for HTML career/search pages.

This source is deliberately best-effort. Structured ATS APIs remain the
primary path, but some requested sources (Monster, Jooble, public career pages,
and Google/LinkedIn search pages) expose only HTML that can change or be
protected by anti-bot systems. Scrapling gives us a stronger fetcher/parser
without letting one blocked page fail the full 6-hour scraper.
"""

from __future__ import annotations

import asyncio
import logging
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import quote_plus, urljoin, urlparse

from bs4 import BeautifulSoup

from app.config import settings
from app.etl.scrapegraph_targets import CAREER_PAGES, google_jobs_url, linkedin_search_url
from app.job_taxonomy import all_role_names, all_taxonomy_scrape_search_terms, categorize
from app.models.job import JobCategory, JobPost, JobSource, VisaBadges
from app.services.global_visa_rules import COUNTRY_RULES, TARGET_COUNTRIES
from app.services.visa_classifier import classify_job
from app.utils.deduplication import generate_content_hash, generate_job_id
from app.utils.job_quality import is_probably_job_search_page

logger = logging.getLogger(__name__)

BACKEND_DIR = Path(__file__).resolve().parents[2]
H1B_EXCEL_FILE = BACKEND_DIR / "H1b_US_DataLIst.xlsx"

DEFAULT_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
}

POPULAR_ROLE_SEEDS = (
    "Software Engineer",
    "Data Engineer",
    "Data Scientist",
    "Machine Learning Engineer",
    "Business Analyst",
    "Financial Analyst",
    "Cybersecurity Analyst",
    "Product Manager (Tech)",
    "Mechanical Engineer",
    "Clinical Research Associate",
)


def _global_location_names() -> list[str]:
    names: list[str] = []
    for code in TARGET_COUNTRIES:
        rule = COUNTRY_RULES.get(code)
        names.append(rule.name if rule else code)
    return names


def _balanced_locations(locations: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for loc in [*locations, *_global_location_names()]:
        clean = _clean(loc)
        key = clean.lower()
        if clean and key not in seen:
            seen.add(key)
            ordered.append(clean)
    return ordered


def _clean(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "")).strip()


def _source_for_kind(kind: str) -> JobSource:
    if kind == "glassdoor":
        return JobSource.GLASSDOOR
    if kind == "ziprecruiter":
        return JobSource.ZIPRECRUITER
    if kind == "monster":
        return JobSource.MONSTER
    if kind == "jooble":
        return JobSource.JOOBLE
    if kind == "linkedin":
        return JobSource.LINKEDIN
    if kind == "google_jobs":
        return JobSource.GOOGLE
    return JobSource.SCRAPLING_DISCOVERY


def _html_from_page(page: Any) -> str:
    for attr in ("html", "text", "body", "content"):
        value = getattr(page, attr, None)
        if callable(value):
            try:
                value = value()
            except TypeError:
                pass
        if isinstance(value, bytes):
            return value.decode("utf-8", errors="ignore")
        if isinstance(value, str) and "<" in value:
            return value
    return str(page or "")


def _looks_like_job_title(title: str) -> bool:
    if not title or len(title) < 4 or len(title) > 140:
        return False
    low = title.lower()
    bad = (
        "privacy", "cookie", "terms", "sign in", "login", "create alert",
        "learn more", "view all", "saved jobs", "job alert", "similar jobs",
        "skip to", "home page", "career advice",
    )
    if any(word in low for word in bad):
        return False
    cat, role = categorize(title)
    if role != "Other":
        return True
    role_terms = {r.lower() for r in all_role_names()}
    return any(term in low for term in role_terms)


def _infer_company(anchor_text: str, target: dict[str, Any], host: str) -> str:
    company = _clean(target.get("company"))
    if company:
        return company
    text = _clean(anchor_text)
    pieces = re.split(r"\s+[-|@]\s+", text)
    if len(pieces) >= 2 and len(pieces[-1]) <= 80:
        return pieces[-1]
    host = host.replace("www.", "")
    return host.split(".")[0].replace("-", " ").title()


def _infer_location(text: str, target: dict[str, Any]) -> str:
    explicit = _clean(target.get("location"))
    if explicit:
        return explicit
    low = text.lower()
    if "remote" in low:
        return "Remote, United States"
    for state in ("CA", "NY", "TX", "WA", "MA", "IL", "NJ", "FL", "GA", "VA", "NC"):
        if re.search(rf"\b{state}\b", text):
            return f"United States ({state})"
    return "United States"


def _html_to_jobs(html: str, target: dict[str, Any]) -> list[JobPost]:
    from app.services.job_filters import is_target_country_scope, is_target_experience, parse_years

    if not html:
        return []
    soup = BeautifulSoup(html, "html.parser")
    base_url = str(target.get("url") or "")
    host = urlparse(base_url).netloc
    source = _source_for_kind(str(target.get("kind") or "scrapling"))

    jobs: list[JobPost] = []
    seen_links: set[str] = set()
    anchors = soup.select("a[href]")
    for anchor in anchors[:1000]:
        title = _clean(anchor.get_text(" ", strip=True))
        href = _clean(anchor.get("href"))
        if not href or not _looks_like_job_title(title):
            continue
        job_url = urljoin(base_url, href)
        if job_url in seen_links:
            continue
        seen_links.add(job_url)

        parent_text = _clean(anchor.parent.get_text(" ", strip=True) if anchor.parent else title)
        company = _infer_company(parent_text or title, target, host)
        if is_probably_job_search_page(title, company, parent_text, source.value):
            continue
        location = _infer_location(parent_text, target)
        if not is_target_country_scope(f"{location} {target.get('location') or ''} {title}"):
            continue
        ymin, ymax = parse_years(f"{title}\n{parent_text}")
        if not is_target_experience(title, ymin, ymax, max_years=10):
            continue

        visa_result = classify_job(title=title, company=company, description=parent_text)
        cat, role = categorize(title)
        extra = {
            "scrapling_discovery": True,
            "target_kind": target.get("kind"),
            "target_url": base_url,
            "target_query": target.get("query"),
            "taxonomy_category": cat if cat != "Other" else None,
            "taxonomy_role": role if role != "Other" else None,
            "years_min": ymin,
            "years_max": ymax,
            "target_experience": True,
            "target_experience_max_years": 10,
        }
        jobs.append(JobPost(
            id=generate_job_id(title, company, location or job_url),
            title=title,
            company=company,
            location=location,
            description=parent_text[:4000],
            job_url=job_url,
            category=JobCategory.OTHER,
            source=source,
            source_job_id=f"{source.value}:{generate_content_hash(title, company, job_url)}",
            content_hash=generate_content_hash(title, company, location or job_url),
            scraped_at=datetime.utcnow(),
            visa=VisaBadges(
                visa_opt=visa_result.visa_opt,
                visa_stem_opt=visa_result.visa_stem_opt,
                visa_h1b=visa_result.visa_h1b,
                h1b_verified=visa_result.h1b_verified,
                no_sponsorship=visa_result.should_discard,
                visa_score=visa_result.score,
            ),
            extra_metadata={k: v for k, v in extra.items() if v not in (None, "", [], {})},
        ))
    return jobs


def _fetch_html(url: str) -> str:
    """Fetch with Scrapling first; fall back to httpx if the package/API changes."""
    try:
        from scrapling.fetchers import Fetcher

        try:
            page = Fetcher.get(url, stealthy_headers=True, timeout=45)
        except TypeError:
            page = Fetcher.get(url)
        return _html_from_page(page)
    except Exception as exc:
        logger.debug("Scrapling fetch failed for %s: %s", url, exc)

    try:
        import httpx

        with httpx.Client(timeout=35.0, headers=DEFAULT_HEADERS, follow_redirects=True) as client:
            response = client.get(url)
            response.raise_for_status()
            return response.text
    except Exception as exc:
        logger.debug("Fallback fetch failed for %s: %s", url, exc)
        return ""


def _extra_career_urls() -> list[str]:
    return [url.strip() for url in settings.scrapegraph_career_pages.split(",") if url.strip()]


def _top_h1b_excel_companies(limit: int) -> list[str]:
    if limit <= 0 or not H1B_EXCEL_FILE.exists():
        return []
    try:
        import openpyxl

        wb = openpyxl.load_workbook(str(H1B_EXCEL_FILE), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        totals: dict[str, int] = {}
        first = True
        for row in ws.iter_rows(values_only=True):
            if first:
                first = False
                continue
            employer = _clean(row[2] if len(row) > 2 else "")
            if not employer:
                continue
            total = 0
            for idx in range(8, min(len(row), 20)):
                try:
                    total += int(float(str(row[idx] or 0).replace(",", "")))
                except Exception:
                    continue
            totals[employer] = totals.get(employer, 0) + total
        return [name for name, _ in sorted(totals.items(), key=lambda item: item[1], reverse=True)[:limit]]
    except Exception as exc:
        logger.info("Scrapling discovery: H1B Excel company load skipped: %s", exc)
        return []


def _top_imported_visa_sponsor_companies(limit: int) -> list[str]:
    if limit <= 0:
        return []
    try:
        from sqlalchemy import text

        from app.db.postgres import PostgresClient

        client = PostgresClient()
        with client.session() as db:
            rows = db.execute(
                text(
                    """
                    select employer_name
                      from visa_sponsors
                     where employer_name is not null
                     group by employer_name
                     order by max(total_petitions) desc, max(approvals) desc, employer_name asc
                     limit :limit
                    """
                ),
                {"limit": limit},
            ).scalars().all()
        return [_clean(row) for row in rows if _clean(row)]
    except Exception as exc:
        logger.info("Scrapling discovery: imported visa sponsor company load skipped: %s", exc)
        return []


def monster_search_url(search_term: str, location: str = "United States") -> str:
    return (
        "https://www.monster.com/jobs/search"
        f"?q={quote_plus(search_term)}&where={quote_plus(location)}"
    )


def jooble_search_url(search_term: str, location: str = "United States") -> str:
    return (
        "https://jooble.org/SearchResult"
        f"?ukw={quote_plus(search_term)}&rgns={quote_plus(location)}"
    )


def ziprecruiter_search_url(search_term: str, location: str = "United States") -> str:
    return (
        "https://www.ziprecruiter.com/jobs-search"
        f"?search={quote_plus(search_term)}&location={quote_plus(location)}"
    )


def glassdoor_search_url(search_term: str, location: str = "United States") -> str:
    return (
        "https://www.glassdoor.com/Job/jobs.htm"
        f"?sc.keyword={quote_plus(search_term)}&locKeyword={quote_plus(location)}"
    )


def build_scrapling_targets(
    *,
    search_terms: Iterable[str] = (),
    locations: Iterable[str] = ("United States",),
    include_glassdoor: bool = False,
    include_ziprecruiter: bool = False,
    include_monster: bool = True,
    include_jooble: bool = True,
    include_linkedin: bool = False,
    include_discovery: bool = True,
    include_search_pages: bool = False,
) -> list[dict[str, Any]]:
    """Build bounded Scrapling targets from roles, H1B sponsors, and career pages."""
    glassdoor_targets: list[dict[str, Any]] = []
    ziprecruiter_targets: list[dict[str, Any]] = []
    monster_targets: list[dict[str, Any]] = []
    jooble_targets: list[dict[str, Any]] = []
    linkedin_targets: list[dict[str, Any]] = []
    discovery_targets: list[dict[str, Any]] = []
    terms = list(search_terms) or all_taxonomy_scrape_search_terms()
    locs = _balanced_locations(locations or ())

    for term in terms:
        for location in locs:
            if include_glassdoor:
                glassdoor_targets.append({"kind": "glassdoor", "url": glassdoor_search_url(term, location), "query": term, "location": location})
            if include_ziprecruiter:
                ziprecruiter_targets.append({"kind": "ziprecruiter", "url": ziprecruiter_search_url(term, location), "query": term, "location": location})
            if include_monster:
                monster_targets.append({"kind": "monster", "url": monster_search_url(term, location), "query": term, "location": location})
            if include_jooble:
                jooble_targets.append({"kind": "jooble", "url": jooble_search_url(term, location), "query": term, "location": location})
            if include_linkedin:
                linkedin_targets.append({"kind": "linkedin", "url": linkedin_search_url(term, location), "query": term, "location": location})

    if include_discovery:
        for entry in CAREER_PAGES:
            discovery_targets.append({"kind": "career_page", "url": entry["url"], "company": entry["company"]})
        for url in _extra_career_urls():
            discovery_targets.append({"kind": "career_page", "url": url})
        if include_search_pages:
            for role in all_role_names():
                discovery_targets.append({"kind": "google_jobs", "url": google_jobs_url(f"{role} OPT H-1B visa sponsor"), "query": role})
                if include_linkedin:
                    for location in locs:
                        discovery_targets.append({"kind": "linkedin", "url": linkedin_search_url(role, location), "query": role, "location": location})
            for company in _top_h1b_excel_companies(settings.scrapling_h1b_excel_company_limit):
                for location in locs:
                    for role in POPULAR_ROLE_SEEDS:
                        query = f"{company} {role} jobs {location}"
                        discovery_targets.append({"kind": "google_jobs", "url": google_jobs_url(query), "query": query, "company": company, "location": location})
            for company in _top_imported_visa_sponsor_companies(settings.scrapling_h1b_excel_company_limit):
                for location in locs:
                    for role in POPULAR_ROLE_SEEDS:
                        query = f"{company} {role} open jobs {location}"
                        discovery_targets.append({"kind": "google_jobs", "url": google_jobs_url(query), "query": query, "company": company, "location": location})

    max_targets = settings.scrapling_discovery_max_targets
    buckets = [bucket for bucket in (glassdoor_targets, ziprecruiter_targets, monster_targets, jooble_targets, linkedin_targets, discovery_targets) if bucket]
    targets: list[dict[str, Any]] = []
    if max_targets <= 0:
        for bucket in buckets:
            targets.extend(bucket)
        return targets

    # Balanced selection: do not let the 533 taxonomy terms consume the whole
    # budget before career-page/H1B discovery gets a turn.
    index = 0
    while len(targets) < max_targets and any(index < len(bucket) for bucket in buckets):
        for bucket in buckets:
            if len(targets) >= max_targets:
                break
            if index < len(bucket):
                targets.append(bucket[index])
        index += 1
    return targets


async def scrape_scrapling_targets(targets: list[dict[str, Any]]) -> list[JobPost]:
    if not settings.scrapling_discovery_enabled or not targets:
        return []
    semaphore = asyncio.Semaphore(settings.scrapling_discovery_concurrency)
    discovered: list[JobPost] = []

    async def _run(target: dict[str, Any]) -> None:
        async with semaphore:
            try:
                html = await asyncio.wait_for(asyncio.to_thread(_fetch_html, str(target["url"])), timeout=75)
                jobs = _html_to_jobs(html, target)
                discovered.extend(jobs)
                logger.info("Scrapling discovery: %s jobs from %s", len(jobs), target.get("url"))
            except Exception as exc:
                logger.info("Scrapling discovery skipped %s: %s", target.get("url"), exc)

    await asyncio.gather(*[_run(target) for target in targets])
    seen: set[str] = set()
    unique: list[JobPost] = []
    for job in discovered:
        if job.content_hash in seen:
            continue
        seen.add(job.content_hash)
        unique.append(job)
    logger.info("Scrapling discovery: %s unique jobs from %s targets", len(unique), len(targets))
    return unique
