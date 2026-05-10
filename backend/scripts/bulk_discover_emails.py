"""
PlaceUp - Bulk Email Discovery (no individual LinkedIn URLs needed)

Workflow per company:
  1. Resolve domain via curated sponsor_domains map (covers top 200+).
  2. Hunter /v2/domain-search?domain=stripe.com → returns up to 100
     (name, position, email, linkedin) records in ONE call.
     Hunter consumes 1 search credit per company, regardless of how many
     contacts come back.
  3. (Optional) For Hunter rows that returned a name but no email, batch
     them into FinalScout /v1/find/professional/bulk to get verified
     emails. Submitted in batches of N, polled until completion.
  4. Aggregate, dedup, fuzzy-match company against the 22,901-row H1B list
     in data/h1b/h1b_us_companies.csv.
  5. Output: 2-sheet xlsx (h1b_company_profiles + regular_company_profiles)
     with columns: name, company, h1b_match, position, role, email,
     linkedin_url, source, confidence.

Usage:
    # Cover top 25 H1B companies on Hunter free tier
    python scripts/bulk_discover_emails.py --top-n 25

    # Cover ALL 22,901 (needs Hunter Pro at $124/mo for 2,500 searches)
    python scripts/bulk_discover_emails.py --all

    # Hunter only (no FinalScout enrichment of name-only rows)
    python scripts/bulk_discover_emails.py --top-n 100 --no-finalscout

    # FinalScout only — discovery via FinalScout /find/professional/bulk
    # (uses common recruiter-name patterns × company; experimental)
    python scripts/bulk_discover_emails.py --top-n 25 --finalscout-only

    # Resumable
    python scripts/bulk_discover_emails.py --top-n 50 --resume

    # Custom output
    python scripts/bulk_discover_emails.py --top-n 25 --output data/exports/bulk_emails.xlsx
"""
from __future__ import annotations

import argparse
import asyncio
import csv
import json
import logging
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from app.config import settings
from app.services.sponsor_domains import best_domain, is_safe_domain, confident_domain
from app.services.hunter_enrichment import domain_search as hunter_domain_search
from app.services.finalscout_enrichment import (
    submit_professional_bulk, submit_linkedin_bulk, collect_bulk_contacts,
)

logging.basicConfig(level=logging.WARNING,
                    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
                    datefmt="%H:%M:%S")
logger = logging.getLogger("placeup.bulk_discover")
logger.setLevel(logging.INFO)


COMMON_RECRUITER_NAMES = [
    # 20 most-common (firstname, lastname) pairs of US recruiters per
    # public Bureau of Labor Statistics + LinkedIn industry reports.
    # These are seed pairs for FinalScout /find/professional/bulk discovery
    # when Hunter returns nothing.
    ("Jennifer", "Smith"), ("Michael", "Johnson"), ("Sarah", "Williams"),
    ("David", "Brown"), ("Jessica", "Jones"), ("Christopher", "Garcia"),
    ("Amanda", "Miller"), ("Matthew", "Davis"), ("Ashley", "Rodriguez"),
    ("Joshua", "Martinez"), ("Emily", "Hernandez"), ("Andrew", "Lopez"),
    ("Stephanie", "Gonzalez"), ("Daniel", "Wilson"), ("Nicole", "Anderson"),
    ("Brian", "Thomas"), ("Rachel", "Taylor"), ("Kevin", "Moore"),
    ("Lauren", "Jackson"), ("Justin", "Martin"),
]


def parse_args():
    p = argparse.ArgumentParser(description="Bulk-discover recruiter emails per H1B company")
    p.add_argument("--h1b-csv", default="data/h1b/h1b_us_companies.csv")
    p.add_argument("--top-n", type=int, default=None,
                   help="Process only top N companies (default: ALL companies)")
    p.add_argument("--all", action="store_true",
                   help="Process ALL companies in the H1B list (overrides --top-n)")
    p.add_argument("--limit-per-company", type=int, default=25,
                   help="Max emails to fetch per company (Hunter caps at 100)")
    p.add_argument("--max-hunter-calls", type=int, default=999_999,
                   help="Stop Hunter after N calls (default: unlimited; free tier: 25/mo)")
    p.add_argument("--no-hunter", action="store_true", help="Skip Hunter")
    p.add_argument("--no-finalscout", action="store_true",
                   help="Skip FinalScout enrichment of name-only rows")
    p.add_argument("--finalscout-only", action="store_true",
                   help="Skip Hunter; use FinalScout with seed names per company")
    p.add_argument("--finalscout-batch-size", type=int, default=50,
                   help="Persons per FinalScout bulk submission (max 100)")
    p.add_argument("--seed-names-per-company", type=int, default=5,
                   help="When --finalscout-only, how many seed names per company")
    p.add_argument("--resume", action="store_true",
                   help="Skip companies already in output CSV")
    p.add_argument("--concurrency", type=int, default=2,
                   help="Parallel Hunter calls (low — be polite)")
    p.add_argument("--output", default=None,
                   help="Output xlsx path (default: data/exports/bulk_emails_<ts>.xlsx)")
    p.add_argument("--match-threshold", type=int, default=85,
                   help="Fuzzy company-match threshold (0-100)")
    return p.parse_args()


# ── helpers ──────────────────────────────────────────────────

def _load_companies(csv_path: Path, top_n: int | None, take_all: bool) -> list[dict]:
    rows = []
    with csv_path.open("r", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            rows.append({
                "company": (r.get("employer_name") or "").strip(),
                "petitions": int(r.get("petitions") or 0),
                "city": r.get("city") or "",
                "state": r.get("state") or "",
            })
    rows.sort(key=lambda r: -r["petitions"])
    if take_all or top_n is None:
        return rows
    return rows[:top_n]


def _normalize_company_name(name: str) -> str:
    if not name: return ""
    n = name.upper().strip()
    for s in [" INC", " LLC", " LTD", " LIMITED", " LLP", " LP", " CORPORATION",
              " CORP", " CO", " COMPANY", "., ", ".", " THE", " US", " USA",
              " GLOBAL", " AMERICAS", " AMERICA", " HOLDINGS", " GROUP",
              " ENTERPRISES", " SERVICES", " SOLUTIONS", " TECHNOLOGY",
              " TECHNOLOGIES", " CONSULTING"]:
        n = n.replace(s, " ")
    n = " ".join(n.split())
    return n.rstrip(",.").strip()


def _fuzzy(a: str, b: str) -> int:
    from difflib import SequenceMatcher
    if not a or not b: return 0
    na, nb = _normalize_company_name(a), _normalize_company_name(b)
    if na == nb: return 100
    a_first = na.split()[0] if na.split() else ""
    b_first = nb.split()[0] if nb.split() else ""
    if a_first and a_first == b_first and (na in nb or nb in na):
        return 95
    return int(SequenceMatcher(None, na, nb).ratio() * 100)


def _classify(company: str, h1b_index: dict, threshold: int) -> tuple[bool, str, int]:
    if not company: return False, "", 0
    norm = _normalize_company_name(company)
    if norm in h1b_index:
        return True, h1b_index[norm]["employer_name"], 100
    first = norm.split()[0] if norm.split() else ""
    candidates = [(k, v) for k, v in h1b_index.items() if k.startswith(first)]
    if not candidates and first:
        candidates = [(k, v) for k, v in h1b_index.items() if first in k]
    best = ("", 0, None)
    for k, v in candidates[:200]:
        s = _fuzzy(norm, k)
        if s > best[1]:
            best = (k, s, v)
    if best[1] >= threshold:
        return True, best[2]["employer_name"], best[1]
    return False, "", best[1]


def _load_h1b_index(csv_path: Path) -> dict[str, dict]:
    out = {}
    with csv_path.open("r", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            name = (r.get("employer_name") or "").strip()
            if not name: continue
            n = _normalize_company_name(name)
            if n and n not in out:
                out[n] = r
    return out


def _dedup_key(row: dict):
    e = (row.get("email") or "").lower().strip()
    li = (row.get("linkedin_url") or "").lower().strip()
    name = (row.get("full_name") or "").lower().strip()
    co = (row.get("company") or "").lower().strip()
    return (
        f"e|{e}" if e
        else (f"l|{li}" if li else (f"n|{co}|{name}" if name else None))
    )


def _load_resume_set(csv_path: Path) -> set[str]:
    if not csv_path.exists():
        return set()
    seen = set()
    with csv_path.open("r", encoding="utf-8") as f:
        for r in csv.DictReader(f):
            if r.get("company"):
                seen.add(r["company"].lower().strip())
    return seen


# ── core ────────────────────────────────────────────────────

def _linkedin_search_urls_for_company(company_name: str) -> list[dict]:
    """Always emit 5 clickable LinkedIn search URLs per company.
    User opens each in their browser as a logged-in human (zero cost, zero ToS risk).
    """
    from urllib.parse import quote_plus
    role_types = [
        ("recruiter", "Recruiter"),
        ("hiring manager", "Hiring Manager"),
        ("talent acquisition", "Talent Acquisition"),
        ("engineering manager", "Engineering Manager"),
        ("technical recruiter", "Technical Recruiter"),
    ]
    rows = []
    for role_keyword, role_label in role_types:
        keywords = quote_plus(f"{role_keyword} {company_name}")
        url = f"https://www.linkedin.com/search/results/people/?keywords={keywords}&geoUrn=%5B%22103644278%22%5D"
        rows.append({
            "company": company_name, "domain": "",
            "full_name": "", "first_name": "", "last_name": "",
            "title": f"LinkedIn search: {role_label} @ {company_name}",
            "role": "recruiter" if "recruit" in role_keyword else "other",
            "email": "", "linkedin_url": url,
            "source": "linkedin_search_url",
            "confidence": "unknown",
        })
    return rows


async def _hunter_one(company: dict, *, limit: int) -> list[dict]:
    """Always emits 5 LinkedIn search URLs per company.
    PLUS Hunter rows when curated domain is available AND Hunter quota remains.
    """
    rows = _linkedin_search_urls_for_company(company["company"])

    domain = confident_domain(company["company"])
    if not domain:
        return rows  # 5 LinkedIn search URLs, no Hunter

    try:
        contacts = await hunter_domain_search(
            domain=domain, company=company["company"], limit=limit,
        )
    except Exception as exc:
        logger.warning("  %s: Hunter failed (%s) -- still emitting LinkedIn URLs", company["company"], exc)
        return rows

    for c in contacts:
        rows.append({
            "company": company["company"], "domain": domain,
            "full_name": c.full_name, "first_name": c.first_name, "last_name": c.last_name,
            "title": c.title, "role": c.role.value if hasattr(c.role, "value") else "other",
            "email": c.email, "linkedin_url": c.linkedin_url,
            "source": "hunter",
            "confidence": c.confidence.value if hasattr(c.confidence, "value") else "unknown",
        })
    return rows


async def _finalscout_seed_one(company: dict, *, n_seeds: int) -> list[dict]:
    """Build (company, seed_name) pairs for FinalScout /find/professional/bulk."""
    persons = []
    for first, last in COMMON_RECRUITER_NAMES[:n_seeds]:
        persons.append({
            "first_name": first, "last_name": last,
            "company": company["company"],
        })
    return persons


async def _finalscout_enrich_name_only_rows(rows: list[dict], batch_size: int) -> list[dict]:
    """For Hunter rows that have name but no email, run FinalScout bulk to get the email."""
    persons = []
    for r in rows:
        if r.get("email") or not (r.get("first_name") and r.get("last_name")):
            continue
        persons.append({
            "first_name": r["first_name"], "last_name": r["last_name"],
            "company": r.get("company") or r.get("domain") or "",
        })
    persons = list({(p["first_name"], p["last_name"], p["company"]): p for p in persons}.values())
    if not persons:
        return []
    logger.info("FinalScout enriching %s name-only rows", len(persons))

    enriched_rows = []
    for i in range(0, len(persons), batch_size):
        batch = persons[i:i + batch_size]
        task_id = await submit_professional_bulk(batch,
            name=f"PlaceUp bulk discover batch {i // batch_size + 1}")
        if not task_id:
            continue
        contacts = await collect_bulk_contacts(task_id)
        for c in contacts:
            enriched_rows.append({
                "company": c.company, "domain": c.company_domain or "",
                "full_name": c.full_name, "first_name": c.first_name, "last_name": c.last_name,
                "title": c.title, "role": c.role.value if hasattr(c.role, "value") else "other",
                "email": c.email, "linkedin_url": c.linkedin_url,
                "source": "finalscout",
                "confidence": c.confidence.value if hasattr(c.confidence, "value") else "unknown",
            })
    return enriched_rows


def _write_xlsx(matched: list[dict], unmatched: list[dict], output_path: Path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    cols = [
        ("name",         lambda r: r.get("full_name") or ""),
        ("company",      lambda r: r.get("company") or ""),
        ("h1b_match",    lambda r: r.get("h1b_match") or ""),
        ("match_score",  lambda r: r.get("match_score") or ""),
        ("position",     lambda r: r.get("title") or ""),
        ("role",         lambda r: r.get("role") or ""),
        ("email",        lambda r: r.get("email") or ""),
        ("linkedin_url", lambda r: r.get("linkedin_url") or ""),
        ("domain",       lambda r: r.get("domain") or ""),
        ("source",       lambda r: r.get("source") or ""),
        ("confidence",   lambda r: r.get("confidence") or ""),
    ]
    header_font = Font(bold=True, color="FFFFFF")
    fills = {
        "h1b_company_profiles": PatternFill("solid", fgColor="1F77B4"),
        "regular_company_profiles": PatternFill("solid", fgColor="2CA02C"),
    }
    for name, rows in (("h1b_company_profiles", matched),
                      ("regular_company_profiles", unmatched)):
        ws = wb.create_sheet(name)
        for ci, (col, _) in enumerate(cols, 1):
            c = ws.cell(row=1, column=ci, value=col)
            c.font = header_font; c.fill = fills[name]
            c.alignment = Alignment(horizontal="center")
        for ri, r in enumerate(rows, 2):
            for ci, (_, accessor) in enumerate(cols, 1):
                ws.cell(row=ri, column=ci, value=accessor(r))
        for ci, (col, _) in enumerate(cols, 1):
            max_len = max([len(str(r.get(col, "") or "")) for r in rows[:200]] + [len(col)])
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max_len + 2, 60)
        ws.freeze_panes = "A2"
    wb.save(output_path)


async def main() -> int:
    args = parse_args()
    h1b_csv = ROOT / args.h1b_csv if not Path(args.h1b_csv).is_absolute() else Path(args.h1b_csv)
    if not h1b_csv.exists():
        print(f"ERROR: H1B CSV not found: {h1b_csv}")
        print("Run scripts/import_h1b_2024_2026.py first.")
        return 1

    companies = _load_companies(h1b_csv, args.top_n, args.all)
    h1b_index = _load_h1b_index(h1b_csv)

    # Resume state
    out_csv_progress = ROOT / "data/exports/bulk_emails_progress.csv"
    out_csv_progress.parent.mkdir(parents=True, exist_ok=True)
    resume_seen = _load_resume_set(out_csv_progress) if args.resume else set()
    if resume_seen:
        before = len(companies)
        companies = [c for c in companies if c["company"].lower().strip() not in resume_seen]
        print(f"Resume mode: skipping {before - len(companies)} companies already in progress CSV")

    print(f"\nBulk discovery on {len(companies)} companies")
    print(f"  Hunter: {'OFF' if args.no_hunter or args.finalscout_only else f'ON (cap: {args.max_hunter_calls} calls)'}")
    print(f"  FinalScout enrich: {'OFF' if args.no_finalscout else 'ON'}")
    print(f"  FinalScout-only seeds: {'ON' if args.finalscout_only else 'OFF'}")
    print()

    # ── Hunter pass ────────────────────────────────────────
    all_rows: list[dict] = []
    hunter_used = 0
    progress_f = out_csv_progress.open("a", newline="", encoding="utf-8")
    progress_w = csv.writer(progress_f)
    if out_csv_progress.stat().st_size == 0:
        progress_w.writerow(["company", "rows_added", "source"])

    if not (args.no_hunter or args.finalscout_only):
        sem = asyncio.Semaphore(args.concurrency)
        async def _bounded(co):
            async with sem:
                return await _hunter_one(co, limit=args.limit_per_company)

        hunter_429_count = 0
        for i, co in enumerate(companies, 1):
            if hunter_used >= args.max_hunter_calls:
                print(f"  *** Hunter cap of {args.max_hunter_calls} reached, switching to LinkedIn-URL-only mode ***")
                # Continue but skip Hunter (still get 5 LinkedIn URLs per company)
                rows = _linkedin_search_urls_for_company(co["company"])
                all_rows.extend(rows)
                continue
            rows = await _bounded(co)
            # Detect Hunter row vs LinkedIn-URL row to track real Hunter usage
            had_hunter_row = any(r.get("source") == "hunter" for r in rows)
            had_li_only = all(r.get("source") == "linkedin_search_url" for r in rows)
            if had_hunter_row:
                hunter_used += 1
            elif not had_li_only:
                # We attempted Hunter but got 0 results back (often 429 or 400)
                hunter_429_count += 1
                if hunter_429_count >= 5:
                    print(f"  *** Hunter returned 0 emails for 5 consecutive companies — quota likely exhausted ***")
                    args.max_hunter_calls = 0  # force LinkedIn-URL-only for the rest
            all_rows.extend(rows)
            with_email = sum(1 for r in rows if r.get("email"))
            print(f"  [{i:>4}/{len(companies)}] HUNTER  {co['company'][:38]:<38}  "
                  f"rows={len(rows):>3}  emails={with_email:>2}  "
                  f"[hunter remaining: {args.max_hunter_calls - hunter_used}]")
            progress_w.writerow([co["company"], len(rows), "hunter"])
            progress_f.flush()

    # ── FinalScout enrich (name-only rows from Hunter) ─────
    if not args.no_finalscout and all_rows and not args.finalscout_only:
        enriched = await _finalscout_enrich_name_only_rows(all_rows, batch_size=args.finalscout_batch_size)
        all_rows.extend(enriched)

    # ── FinalScout-only seed mode ─────────────────────────
    if args.finalscout_only:
        all_persons = []
        for co in companies:
            persons = await _finalscout_seed_one(co, n_seeds=args.seed_names_per_company)
            all_persons.extend(persons)
        print(f"FinalScout seed mode: submitting {len(all_persons)} (name, company) pairs")
        for i in range(0, len(all_persons), args.finalscout_batch_size):
            batch = all_persons[i:i + args.finalscout_batch_size]
            task_id = await submit_professional_bulk(batch, name=f"PlaceUp seed batch {i // args.finalscout_batch_size + 1}")
            if not task_id:
                continue
            contacts = await collect_bulk_contacts(task_id)
            for c in contacts:
                all_rows.append({
                    "company": c.company, "domain": c.company_domain or "",
                    "full_name": c.full_name, "first_name": c.first_name, "last_name": c.last_name,
                    "title": c.title, "role": c.role.value if hasattr(c.role, "value") else "other",
                    "email": c.email, "linkedin_url": c.linkedin_url,
                    "source": "finalscout",
                    "confidence": c.confidence.value if hasattr(c.confidence, "value") else "unknown",
                })

    progress_f.close()

    # ── Dedup ─────────────────────────────────────────────
    seen_keys = set()
    deduped = []
    for r in all_rows:
        k = _dedup_key(r)
        if not k or k in seen_keys: continue
        seen_keys.add(k); deduped.append(r)

    # ── Classify h1b vs regular ───────────────────────────
    matched, unmatched = [], []
    for r in deduped:
        is_h1b, canonical, score = _classify(r.get("company") or "", h1b_index, args.match_threshold)
        r["h1b_match"] = canonical; r["match_score"] = score
        (matched if is_h1b else unmatched).append(r)

    # ── Write xlsx ────────────────────────────────────────
    out = Path(args.output) if args.output else (
        ROOT / "data/exports" / f"bulk_emails_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    _write_xlsx(matched, unmatched, out)

    print()
    print("=" * 70)
    print("BULK DISCOVERY COMPLETE")
    print("=" * 70)
    print(f"  Companies attempted:       {len(companies)}")
    print(f"  Total rows discovered:     {len(deduped)}  (was {len(all_rows)} pre-dedup)")
    print(f"  Matched H1B sponsor:       {len(matched)}")
    print(f"    - with email:            {sum(1 for r in matched if r.get('email'))}")
    print(f"    - with linkedin URL:     {sum(1 for r in matched if r.get('linkedin_url'))}")
    print(f"  Regular companies:         {len(unmatched)}")
    print(f"    - with email:            {sum(1 for r in unmatched if r.get('email'))}")
    print(f"  Hunter calls used:         {hunter_used}")
    print(f"  Output xlsx:               {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(asyncio.run(main()))
